import openpyxl
import json
import random
import os


#Variables estáticas / condiciones de borde
AMP = 'MATRICES/Matriz AMP.xlsx' #Aquí se va a buscar la matrix ideal del KAM, ya sea local o en una nube ej en S3
KAM = 'MATRICES/Matriz KAM.xlsx' #Aquí se va a buscar la matrix ideal del KAM, ya sea local o en una nube ej en S3
FLSM = 'MATRICES/Matriz FLSM.xlsx' #Aquí se va a buscar la matrix ideal del KAM, ya sea local o en una nube ej en S3
niveles = {'Fundamental' : ['C','D','E','F'], 'Regular' : ['G','H','I','J'], 'Profesional': ['K','L','M','N']}
columna_comportamientos = 'B'
output = []
ideal = False

def iterar_comportamiento(hoja,primer_comportamiento, ultimo_comportamiento, nombre, padre, id, nivel, ROLE, pais=None):

    dato = {'id': id,
           'nivel': nivel, #0,
           'role': ROLE,
           'pais': pais,
           'nombre': nombre, #next.value,
           'padre': padre, #competencia_principal,
           'comportamientos': [
                {
                    'list': [],
                    'scores': []
                }
            ]
          }

    for fila_comportamiento in range(primer_comportamiento, ultimo_comportamiento + 1):

        celda_comportamiento = hoja[columna_comportamientos + str(fila_comportamiento)]
        #Se agrega el comportamiento
        dato['comportamientos'][0]['list'].append(celda_comportamiento.value)
        #Para cada comportamiento ver el score
        if id == 0:
            score = 0
            if nivel == 'Fundamental':
                score = random.randint(0,1)
            elif nivel == 'Regular':
                score = random.randint(1,2)
            elif nivel == 'Profesional':
                score = random.randint(2,3)
            dato['comportamientos'][0]['scores'].append(score)
        else:
            score = random.randint(0,3)
            dato['comportamientos'][0]['scores'].append(score)

    return dato

def main_ideal(output, role, ROLE):


    wb = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__)) + '/' + role)
    competencias_principales = wb.sheetnames
    #print('[LOG1] {0}'.format(competencias_principales))
    for nivel in ['Fundamental', 'Regular', 'Profesional']:
        #print('[LOG2] {0}'.format(nivel))
        for competencia_principal in competencias_principales:

            hoja = wb[competencia_principal]
            #print('[LOG3] {0}'.format(competencia_principal))
            posicion_next = 'A11'

            #posicion_next = 'A' + str(ultimo_comportamiento + 4)
            next = hoja[posicion_next]

            while next.value:

                for cell in hoja.merged_cells.ranges:
                    if posicion_next in cell:
                        bloque = cell

                primer_comportamiento = bloque.bounds[1]
                ultimo_comportamiento = bloque.bounds[3]
                #print('[LOG4] {0}'.format(next.value))
                dato = iterar_comportamiento(hoja, primer_comportamiento, ultimo_comportamiento, next.value, competencia_principal, 0, nivel, ROLE)
                output.append(dato)


                posicion_next = 'A' + str(ultimo_comportamiento + 4)
                next = hoja[posicion_next]

    return output

def main_fake(output, role, ROLE):


    wb = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__)) + '/' + role)
    competencias_principales = wb.sheetnames

    for id in range(1,6):

        nivel = random.choice(['Fundamental', 'Regular', 'Profesional'])
        pais = random.choice(['Chile', 'Argentina', 'Uruguay'])
        for competencia_principal in competencias_principales:

            hoja = wb[competencia_principal]

            posicion_next = 'A11'

            #posicion_next = 'A' + str(ultimo_comportamiento + 4)
            next = hoja[posicion_next]

            while next.value:

                for cell in hoja.merged_cells.ranges:
                    if posicion_next in cell:
                        bloque = cell

                primer_comportamiento = bloque.bounds[1]
                ultimo_comportamiento = bloque.bounds[3]

                dato = iterar_comportamiento(hoja, primer_comportamiento, ultimo_comportamiento, next.value, competencia_principal, id, nivel, ROLE, pais)
                output.append(dato)


                posicion_next = 'A' + str(ultimo_comportamiento + 4)
                next = hoja[posicion_next]

    return output



#1. determinar si rol x es nivel x, # de personas y % de personas
#2. determinar % de personal en nivel x
#3. determinar principales y nivel de Expertiz por rol

amp_ideal = main_ideal([], AMP, 'AMP')
amp_fake = main_fake([], AMP, 'AMP')
kam_ideal = main_ideal([], KAM, 'KAM')
kam_fake = main_fake([], KAM, 'KAM')
flsm_ideal = main_ideal([], FLSM, 'FLSM')
flsm_fake = main_fake([], FLSM, 'FLSM')

output = amp_ideal + amp_fake + kam_ideal + kam_fake + flsm_ideal + flsm_fake
output = json.dumps(output,ensure_ascii=False, indent=4) #quitar el indent
